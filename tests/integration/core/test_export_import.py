"""Route-level integration tests for export (the built-in `export` batch
action, `POST /_api/{key}/action?name=export`) and import
(`POST /_api/{key}/import`, with `?preview=1` for the wizard's preview step).

Uses the TinyDB-backed test infrastructure so the tests are backend-agnostic.
"""

from __future__ import annotations

import csv
import io
import json
import zipfile
from urllib.parse import urlencode

import openpyxl
import pytest
from starlette.applications import Starlette
from starlette_admin import BaseAdmin, FileField, FloatField, IntegerField, StringField
from starlette_admin.export import (
    CsvExporter,
    JsonExporter,
    PdfExporter,
    TablibExporter,
)
from starlette_admin.importers import CsvImporter, JsonImporter, TablibImporter

from tests.integration.core.tinydb_model_view import TinydbBaseModel, TinydbModelView
from tests.utils import CsrfTestClient

# ── Test model & views ────────────────────────────────────────────────────────


class Product(TinydbBaseModel):
    name: str
    price: float = 0.0
    sku: str = ""


class ProductView(TinydbModelView):
    key = "product"
    fields = [
        IntegerField("id"),
        StringField("name"),
        FloatField("price"),
        StringField("sku"),
    ]
    exporters = [
        CsvExporter(escape_formulas=True),
        TablibExporter("xlsx"),
        JsonExporter(),
        PdfExporter(),
    ]
    importers = [CsvImporter(), TablibImporter("xlsx"), JsonImporter()]
    searchable_fields = ("name", "sku")

    def can_import(self, request):
        return True


class PaginatedProductView(TinydbModelView):
    """ProductView clone with a small page size, for export scope=page tests."""

    key = "paginated-product"
    fields = [
        IntegerField("id"),
        StringField("name"),
        FloatField("price"),
        StringField("sku"),
    ]
    exporters = [CsvExporter()]
    page_size = 1
    page_size_options = [1, 2, 3, -1]


class ExportOnlyView(TinydbModelView):
    """Export enabled, import disabled (default)."""

    key = "export-only"
    fields = [IntegerField("id"), StringField("name"), FloatField("price")]
    exporters = [CsvExporter()]

    def can_import(self, request):
        return False


class NoImportExportView(TinydbModelView):
    """Export explicitly disabled."""

    key = "no-export"
    fields = [IntegerField("id"), StringField("name")]

    def can_export(self, request):
        return False

    def can_import(self, request):
        return False


class ExcludeExportFieldsView(TinydbModelView):
    """View that hides the 'sku' field from exports."""

    key = "exclude-export"
    fields = [
        IntegerField("id"),
        StringField("name"),
        FloatField("price"),
        StringField("sku"),
    ]
    exporters = [CsvExporter(), JsonExporter()]
    importers = [CsvImporter()]
    exclude_fields_from_export = ["sku"]
    searchable_fields = ("name",)

    def can_import(self, request):
        return True


class ExcludeImportFieldsView(TinydbModelView):
    """View that ignores the 'price' field during imports."""

    key = "exclude-import"
    fields = [
        IntegerField("id"),
        StringField("name"),
        FloatField("price"),
        StringField("sku"),
    ]
    exporters = [CsvExporter()]
    importers = [CsvImporter()]
    exclude_fields_from_import = ["price"]
    searchable_fields = ("name",)

    def can_import(self, request):
        return True


class Attachment(TinydbBaseModel):
    name: str = ""
    document: dict | None = None


class AttachmentView(TinydbModelView):
    """View with a storage-backed file field for export ZIP / import-exclusion tests."""

    key = "attachment"
    fields = [
        IntegerField("id"),
        StringField("name"),
        FileField("document", storage=None, upload_folder="docs/"),
    ]
    searchable_fields = ("name",)

    def can_import(self, request):
        return True


# ── Fixtures ──────────────────────────────────────────────────────────────────


@pytest.fixture()
def client_with_data():
    """Populated ProductView + TestClient."""
    ProductView._db.truncate()
    ProductView._db.insert(
        Product(name="Widget", price=9.99, sku="W-001").to_tinydb_doc()
    )
    ProductView._db.insert(
        Product(name="Gadget", price=19.99, sku="G-002").to_tinydb_doc()
    )
    ProductView._db.insert(
        Product(name="Doohickey", price=4.99, sku="D-003").to_tinydb_doc()
    )

    app = Starlette()
    admin = BaseAdmin()
    admin.add_view(ProductView(Product))
    admin.mount_to(app)
    return CsrfTestClient(app, raise_server_exceptions=True)


@pytest.fixture()
def client_with_paginated_data():
    """Populated PaginatedProductView (page_size=1) + TestClient."""
    PaginatedProductView._db.truncate()
    PaginatedProductView._db.insert(
        Product(name="Widget", price=9.99, sku="W-001").to_tinydb_doc()
    )
    PaginatedProductView._db.insert(
        Product(name="Gadget", price=19.99, sku="G-002").to_tinydb_doc()
    )
    PaginatedProductView._db.insert(
        Product(name="Doohickey", price=4.99, sku="D-003").to_tinydb_doc()
    )

    app = Starlette()
    admin = BaseAdmin()
    admin.add_view(PaginatedProductView(Product))
    admin.mount_to(app)
    return CsrfTestClient(app, raise_server_exceptions=True)


@pytest.fixture()
def export_only_client():
    ExportOnlyView._db.truncate()
    ExportOnlyView._db.insert(Product(name="Alpha", price=1.0).to_tinydb_doc())

    app = Starlette()
    admin = BaseAdmin()
    admin.add_view(ExportOnlyView(Product))
    admin.mount_to(app)
    return CsrfTestClient(app, raise_server_exceptions=True)


@pytest.fixture()
def no_export_client():
    NoImportExportView._db.truncate()
    NoImportExportView._db.insert(Product(name="Beta", price=2.0).to_tinydb_doc())

    app = Starlette()
    admin = BaseAdmin()
    admin.add_view(NoImportExportView(Product))
    admin.mount_to(app)
    return CsrfTestClient(app, raise_server_exceptions=True)


@pytest.fixture()
def exclude_export_client():
    ExcludeExportFieldsView._db.truncate()
    ExcludeExportFieldsView._db.insert(
        Product(name="Widget", price=9.99, sku="W-001").to_tinydb_doc()
    )
    app = Starlette()
    admin = BaseAdmin()
    admin.add_view(ExcludeExportFieldsView(Product))
    admin.mount_to(app)
    return CsrfTestClient(app, raise_server_exceptions=True)


@pytest.fixture()
def exclude_import_client():
    ExcludeImportFieldsView._db.truncate()
    app = Starlette()
    admin = BaseAdmin()
    admin.add_view(ExcludeImportFieldsView(Product))
    admin.mount_to(app)
    return CsrfTestClient(app, raise_server_exceptions=True)


@pytest.fixture()
def attachment_client(storage):
    """AttachmentView + TestClient, patched with the parametrized storage."""
    document_field = next(f for f in AttachmentView.fields if f.name == "document")
    document_field.storage = storage

    AttachmentView._db.truncate()
    app = Starlette()
    admin = BaseAdmin()
    admin.add_view(AttachmentView(Attachment))
    admin.mount_to(app)
    client = CsrfTestClient(app, raise_server_exceptions=True)
    yield client, storage
    document_field.storage = None


# ── Helpers ───────────────────────────────────────────────────────────────────


def _file_tuple(filename: str, content: bytes, content_type: str) -> tuple:
    """Return a (filename, BytesIO, content_type) tuple for httpx multipart uploads."""
    return (filename, io.BytesIO(content), content_type)


def _csv_upload(rows: list[list[str]]) -> tuple:
    buf = io.StringIO()
    csv.writer(buf).writerows(rows)
    return _file_tuple("data.csv", buf.getvalue().encode(), "text/csv")


def _excel_upload(headers: list[str], *rows: list) -> tuple:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return _file_tuple(
        "data.xlsx",
        buf.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _json_upload(data: list[dict]) -> tuple:
    return _file_tuple("data.json", json.dumps(data).encode(), "application/json")


def _export_url(
    key: str,
    *,
    pks: list | None = None,
    select_all: bool = False,
    list_query: str = "",
    origin: str | None = None,
) -> str:
    """Build the `POST .../action?name=export` URL, with the row-selection
    state carried as query params exactly like `list.js`'s `ActionManager`
    does client side."""
    params = [("name", "export")]
    if list_query:
        params.append(("_list_query", list_query))
    if origin:
        params.append(("_origin", origin))
    if select_all:
        params.append(("all", "1"))
    elif pks:
        params.extend(("pks", str(pk)) for pk in pks)
    return f"/admin/_api/{key}/action?{urlencode(params)}"


def _export(
    client,
    key: str,
    *,
    scope: str = "page",
    fmt: str = "csv",
    filename: str | None = None,
    fields: list[str] | None = None,
    pks: list | None = None,
    select_all: bool = False,
    list_query: str = "",
    origin: str | None = None,
    follow_redirects: bool = True,
):
    data: dict = {"scope": scope, "format": fmt}
    if filename is not None:
        data["filename"] = filename
    if fields is not None:
        data["fields"] = fields
    return client.post(
        _export_url(
            key, pks=pks, select_all=select_all, list_query=list_query, origin=origin
        ),
        data=data,
        follow_redirects=follow_redirects,
    )


# ── Export: happy paths ───────────────────────────────────────────────────────


class TestExportHappyPath:
    def test_csv_export_status_and_content_type(self, client_with_data):
        r = _export(client_with_data, "product", fmt="csv")
        assert r.status_code == 200
        assert "text/csv" in r.headers["content-type"]

    def test_csv_export_contains_all_rows(self, client_with_data):
        r = _export(client_with_data, "product", fmt="csv")
        text = r.text
        assert "Widget" in text
        assert "Gadget" in text
        assert "Doohickey" in text

    def test_csv_export_escapes_formula_injection(self, client_with_data):
        ProductView._db.insert(
            Product(name="=cmd|'/c calc'!A1", price=0.0).to_tinydb_doc()
        )
        r = _export(client_with_data, "product", fmt="csv")
        reader = csv.reader(io.StringIO(r.text))
        rows = list(reader)[1:]  # skip header
        names = [row[1] for row in rows]
        assert "'=cmd|'/c calc'!A1" in names

    def test_csv_export_has_header_row(self, client_with_data):
        r = _export(client_with_data, "product", fmt="csv")
        reader = csv.reader(io.StringIO(r.text))
        header = next(reader)
        assert "Name" in header
        assert "Price" in header

    def test_excel_export_status_and_content_type(self, client_with_data):
        r = _export(client_with_data, "product", fmt="xlsx")
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers["content-type"]

    def test_excel_export_contains_all_rows(self, client_with_data):
        r = _export(client_with_data, "product", fmt="xlsx")
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        data = list(ws.iter_rows(values_only=True))
        all_values = [str(cell) for row in data for cell in row if cell is not None]
        assert "Widget" in all_values
        assert "Gadget" in all_values

    def test_json_export_status_and_content_type(self, client_with_data):
        r = _export(client_with_data, "product", fmt="json")
        assert r.status_code == 200
        assert "application/json" in r.headers["content-type"]

    def test_json_export_contains_all_rows(self, client_with_data):
        r = _export(client_with_data, "product", fmt="json")
        data = r.json()
        names = [row.get("Name") or row.get("name") for row in data]
        assert "Widget" in names
        assert "Gadget" in names

    def test_pdf_export_returns_pdf_bytes(self, client_with_data):
        r = _export(client_with_data, "product", fmt="pdf")
        assert r.status_code == 200
        assert r.content[:4] == b"%PDF"

    def test_export_content_disposition_filename(self, client_with_data):
        r = _export(client_with_data, "product", fmt="csv")
        assert "product" in r.headers.get("content-disposition", "")

    def test_export_custom_filename(self, client_with_data):
        r = _export(client_with_data, "product", fmt="csv", filename="my-export")
        assert "my-export" in r.headers.get("content-disposition", "")

    def test_export_respects_search_param(self, client_with_data):
        r = _export(client_with_data, "product", fmt="csv", list_query="q=widget")
        reader = csv.reader(io.StringIO(r.text))
        rows = list(reader)[1:]  # skip header
        names = [row[1] for row in rows if row]  # id, name, price, sku
        assert all("widget" in n.lower() for n in names)

    def test_export_scope_page_limits_to_current_page(self, client_with_paginated_data):
        r = _export(
            client_with_paginated_data,
            "paginated-product",
            fmt="csv",
            scope="page",
            list_query="page=1&page_size=1",
        )
        reader = csv.reader(io.StringIO(r.text))
        rows = list(reader)[1:]  # skip header
        assert len(rows) == 1

    def test_export_scope_page_second_page_returns_different_row(
        self, client_with_paginated_data
    ):
        page1 = _export(
            client_with_paginated_data,
            "paginated-product",
            fmt="csv",
            scope="page",
            list_query="page=1&page_size=1",
        )
        page2 = _export(
            client_with_paginated_data,
            "paginated-product",
            fmt="csv",
            scope="page",
            list_query="page=2&page_size=1",
        )
        row1 = list(csv.reader(io.StringIO(page1.text)))[1]
        row2 = list(csv.reader(io.StringIO(page2.text)))[1]
        assert row1 != row2

    def test_export_scope_selected_explicit_pks(self, client_with_data):
        pk = ProductView._all_pks()[0]
        r = _export(client_with_data, "product", fmt="csv", scope="selected", pks=[pk])
        reader = csv.reader(io.StringIO(r.text))
        rows = list(reader)[1:]
        assert len(rows) == 1

    def test_export_scope_selected_select_all_matching(self, client_with_data):
        r = _export(
            client_with_data,
            "product",
            fmt="csv",
            scope="selected",
            select_all=True,
            list_query="q=Widget",
        )
        reader = csv.reader(io.StringIO(r.text))
        rows = list(reader)[1:]
        assert len(rows) == 1
        assert "Widget" in rows[0][1]

    def test_export_field_subset(self, client_with_data):
        r = _export(client_with_data, "product", fmt="csv", fields=["name"])
        reader = csv.reader(io.StringIO(r.text))
        header = next(reader)
        assert header == ["Name"]

    def test_export_empty_fields_selection_falls_back_to_all(self, client_with_data):
        r = _export(client_with_data, "product", fmt="csv", fields=[])
        reader = csv.reader(io.StringIO(r.text))
        header = next(reader)
        assert "Name" in header
        assert "Price" in header


# ── Export: error paths ───────────────────────────────────────────────────────


class TestExportErrors:
    def test_unknown_format_flashes_error(self, client_with_data):
        r = _export(client_with_data, "product", fmt="docx", follow_redirects=False)
        assert r.status_code == 303

        r = client_with_data.get(r.headers["location"], follow_redirects=True)
        assert r.status_code == 200
        assert "Unknown export format" in r.text

    def test_disabled_format_flashes_error(self, export_only_client):
        # PDF is not in ExportOnlyView.exporters
        r = _export(
            export_only_client, "export-only", fmt="pdf", follow_redirects=False
        )
        assert r.status_code == 303

        r = export_only_client.get(r.headers["location"], follow_redirects=True)
        assert r.status_code == 200
        assert "Unknown export format" in r.text

    def test_error_redirect_honors_origin(self, client_with_data):
        origin = "/admin/product/list?q=widget&sort=name__asc"
        r = _export(
            client_with_data,
            "product",
            fmt="docx",
            origin=origin,
            follow_redirects=False,
        )
        assert r.status_code == 303
        assert r.headers["location"] == origin

    def test_no_permission_returns_400(self, no_export_client):
        # is_action_allowed() -> can_export() is False, so handle_action
        # raises ActionFailed, which the action route reports as 400 JSON
        # (not the export action's own redirect-with-flash path, since the
        # action is rejected before export_action ever runs).
        r = _export(no_export_client, "no-export", fmt="csv", follow_redirects=False)
        assert r.status_code == 400

    def test_unknown_key_returns_404(self, client_with_data):
        r = _export(client_with_data, "nonexistent", fmt="csv")
        assert r.status_code == 404


# ── Import: happy paths ───────────────────────────────────────────────────────


class TestImportHappyPath:
    def setup_method(self, method):
        ProductView._db.truncate()

    def test_csv_import_creates_records(self, client_with_data):
        upload = _csv_upload([["Name", "Price", "Sku"], ["TestItem", "5.0", "T-001"]])
        r = client_with_data.post(
            "/admin/_api/product/import",
            data={"format": "csv"},
            files={"file": upload},
        )
        assert r.status_code == 200
        data = r.json()
        assert data["rows_total"] == 1
        assert data["rows_created"] == 1
        assert not data["has_errors"]

    def test_excel_import_creates_records(self, client_with_data):
        upload = _excel_upload(["Name", "Price", "Sku"], ["ExcelItem", 3.5, "E-001"])
        r = client_with_data.post(
            "/admin/_api/product/import",
            data={"format": "xlsx"},
            files={"file": upload},
        )
        assert r.status_code == 200
        data = r.json()
        assert data["rows_total"] == 1
        assert data["rows_created"] == 1

    def test_json_import_creates_records(self, client_with_data):
        payload = [{"Name": "JsonItem", "Price": "2.0", "Sku": "J-001"}]
        upload = _json_upload(payload)
        r = client_with_data.post(
            "/admin/_api/product/import",
            data={"format": "json"},
            files={"file": upload},
        )
        assert r.status_code == 200
        data = r.json()
        assert data["rows_total"] == 1
        assert data["rows_created"] == 1

    def test_preview_does_not_persist(self, client_with_data):
        upload = _csv_upload([["Name", "Price", "Sku"], ["DryItem", "1.0", "D-001"]])
        r = client_with_data.post(
            "/admin/_api/product/import?preview=1",
            data={"format": "csv"},
            files={"file": upload},
        )
        assert r.status_code == 200
        data = r.json()
        assert data["rows_total"] == 1
        assert data["rows_new"] == 1
        # Verify no new record was written (fixture pre-seeds 3 records)
        assert len(ProductView._db.all()) == 3

    def test_preview_returns_header_mapping(self, client_with_data):
        upload = _csv_upload([["Name", "Price", "Sku"], ["Item", "1.0", "S-001"]])
        r = client_with_data.post(
            "/admin/_api/product/import?preview=1",
            data={"format": "csv"},
            files={"file": upload},
        )
        data = r.json()
        assert data["headers"] == ["Name", "Price", "Sku"]
        mapped = {m["header"]: m["field"] for m in data["mapping"]}
        assert mapped["Name"] == "name"
        assert mapped["Price"] == "price"
        assert data["unmatched_headers"] == []

    def test_preview_reports_unmatched_headers(self, client_with_data):
        upload = _csv_upload([["Name", "Extra"], ["Item", "Should be ignored"]])
        r = client_with_data.post(
            "/admin/_api/product/import?preview=1",
            data={"format": "csv"},
            files={"file": upload},
        )
        data = r.json()
        assert data["unmatched_headers"] == ["Extra"]

    def test_preview_sample_has_first_rows_with_status(self, client_with_data):
        upload = _csv_upload([["Name", "Price", "Sku"], ["Item", "1.0", "S-001"]])
        r = client_with_data.post(
            "/admin/_api/product/import?preview=1",
            data={"format": "csv"},
            files={"file": upload},
        )
        data = r.json()
        assert len(data["sample"]) == 1
        assert data["sample"][0]["status"] == "new"

    def test_import_multiple_rows(self, client_with_data):
        upload = _csv_upload(
            [
                ["Name", "Price", "Sku"],
                ["A", "1.0", "A-001"],
                ["B", "2.0", "B-002"],
                ["C", "3.0", "C-003"],
            ]
        )
        r = client_with_data.post(
            "/admin/_api/product/import",
            data={"format": "csv"},
            files={"file": upload},
        )
        assert r.status_code == 200
        assert r.json()["rows_total"] == 3
        assert r.json()["rows_created"] == 3

    def test_import_empty_file_creates_zero_rows(self, client_with_data):
        upload = _csv_upload([["Name", "Price", "Sku"]])
        r = client_with_data.post(
            "/admin/_api/product/import",
            data={"format": "csv"},
            files={"file": upload},
        )
        assert r.status_code == 200
        data = r.json()
        assert data["rows_total"] == 0
        assert data["rows_created"] == 0

    def test_import_returns_errors_for_bad_rows(self, client_with_data):
        """A row that fails view.create() ends up in result.errors."""
        # Price field will receive a non-numeric value → create() raises
        upload = _csv_upload(
            [["Name", "Price", "Sku"], ["BadItem", "not-a-number", "X-001"]]
        )
        r = client_with_data.post(
            "/admin/_api/product/import",
            data={"format": "csv"},
            files={"file": upload},
        )
        assert r.status_code == 200
        data = r.json()
        assert data["rows_total"] == 1
        assert data["rows_created"] == 0
        assert data["has_errors"]
        assert len(data["errors"]) == 1
        assert data["errors"][0]["row"] == 1

    def test_deselecting_pk_field_lets_backend_auto_generate(self, client_with_data):
        """Omitting "id" from `fields` is the field-selection equivalent of
        the old `skip_pk=1`: the column is ignored, and TinyDB assigns a
        fresh id instead of reusing the uploaded one."""
        upload = _csv_upload(
            [
                ["Id", "Name", "Price", "Sku", "Description"],
                ["999", "HeaderItem", "2.5", "H-001", "Should be ignored"],
            ]
        )
        r = client_with_data.post(
            "/admin/_api/product/import",
            data={"format": "csv", "fields": ["name", "price", "sku"]},
            files={"file": upload},
        )
        assert r.status_code == 200
        data = r.json()
        assert data["rows_total"] == 1
        assert data["rows_created"] == 1
        assert not data["has_errors"]

        record = next(
            (
                ProductView._get(pk)
                for pk in ProductView._all_pks()
                if ProductView._get(pk).name == "HeaderItem"
            ),
            None,
        )
        assert record is not None
        assert record.id != 999

    def test_upsert_updates_existing_record(self, client_with_data):
        pk = ProductView._all_pks()[0]
        original = ProductView._get(pk)
        upload = _csv_upload(
            [["Id", "Name", "Price", "Sku"], [str(pk), "Renamed", "1.23", "R-001"]]
        )
        r = client_with_data.post(
            "/admin/_api/product/import",
            data={"format": "csv", "update_existing": "1"},
            files={"file": upload},
        )
        assert r.status_code == 200
        data = r.json()
        assert data["rows_total"] == 1
        assert data["rows_created"] == 0
        assert data["rows_updated"] == 1

        updated = ProductView._get(pk)
        assert updated.name == "Renamed"
        assert updated.name != original.name

    def test_upsert_creates_when_pk_does_not_match(self, client_with_data):
        upload = _csv_upload(
            [["Id", "Name", "Price", "Sku"], ["999999", "NewOne", "1.0", "N-001"]]
        )
        r = client_with_data.post(
            "/admin/_api/product/import",
            data={"format": "csv", "update_existing": "1"},
            files={"file": upload},
        )
        assert r.status_code == 200
        data = r.json()
        assert data["rows_created"] == 1
        assert data["rows_updated"] == 0

    def test_without_upsert_pk_match_still_creates(self, client_with_data):
        """update_existing defaults to off: a pk collision is not upserted."""
        pk = ProductView._all_pks()[0]
        before = len(ProductView._db.all())
        upload = _csv_upload(
            [["Id", "Name", "Price", "Sku"], [str(pk), "ShouldCreate", "1.0", "C-001"]]
        )
        r = client_with_data.post(
            "/admin/_api/product/import",
            data={"format": "csv"},
            files={"file": upload},
        )
        assert r.status_code == 200
        assert r.json()["rows_created"] == 1
        assert len(ProductView._db.all()) == before + 1

    def test_preview_classifies_new_and_update_rows(self, client_with_data):
        pk = ProductView._all_pks()[0]
        upload = _csv_upload(
            [
                ["Id", "Name", "Price", "Sku"],
                [str(pk), "WillUpdate", "1.0", "U-001"],
                ["", "WillCreate", "2.0", "C-002"],
            ]
        )
        r = client_with_data.post(
            "/admin/_api/product/import?preview=1",
            data={"format": "csv", "update_existing": "1"},
            files={"file": upload},
        )
        data = r.json()
        assert data["rows_new"] == 1
        assert data["rows_updated"] == 1
        statuses = {row["status"] for row in data["sample"]}
        assert statuses == {"new", "update"}
        # Preview must not have actually written the update.
        assert ProductView._get(pk).name != "WillUpdate"


# ── Import: error paths ───────────────────────────────────────────────────────


class TestImportErrors:
    def test_no_file_returns_400(self, client_with_data):
        r = client_with_data.post(
            "/admin/_api/product/import",
            data={"format": "csv"},
        )
        assert r.status_code == 400
        assert "No file" in r.json()["error"]

    def test_unknown_format_returns_400(self, client_with_data):
        upload = _csv_upload([["Name"], ["Widget"]])
        r = client_with_data.post(
            "/admin/_api/product/import",
            data={"format": "docx"},
            files={"file": upload},
        )
        assert r.status_code == 400
        assert "Unknown" in r.json()["error"]

    def test_disabled_format_returns_400(self, client_with_data):
        # PDF is not in ProductView.importers
        upload = ("file", io.BytesIO(b"%PDF"), "application/pdf")
        r = client_with_data.post(
            "/admin/_api/product/import",
            data={"format": "pdf"},
            files={"file": upload},
        )
        assert r.status_code == 400
        assert "Unknown" in r.json()["error"]

    def test_no_permission_returns_403(self, export_only_client):
        # ExportOnlyView.can_import() returns False
        upload = _csv_upload([["Name"], ["Alpha"]])
        r = export_only_client.post(
            "/admin/_api/export-only/import",
            data={"format": "csv"},
            files={"file": upload},
        )
        assert r.status_code == 403

    def test_unknown_key_returns_404(self, client_with_data):
        upload = _csv_upload([["Name"], ["X"]])
        r = client_with_data.post(
            "/admin/_api/nonexistent/import",
            data={"format": "csv"},
            files={"file": upload},
        )
        assert r.status_code == 404

    def test_json_import_invalid_file_returns_400(self, client_with_data):
        upload = _file_tuple("data.json", b"not valid json", "application/json")
        r = client_with_data.post(
            "/admin/_api/product/import",
            data={"format": "json"},
            files={"file": upload},
        )
        assert r.status_code == 400
        assert "error" in r.json()


# ── Storage-backed file field import/export (local + s3) ──────────────────────


class TestFileFieldExportImport:
    @pytest.mark.asyncio
    async def test_csv_export_with_file_field_returns_zip(self, attachment_client):
        client, storage = attachment_client
        r = client.post(
            "/admin/attachment/create",
            data={"name": "Report"},
            files={
                "document": ("report.txt", io.BytesIO(b"report content"), "text/plain")
            },
            follow_redirects=False,
        )
        assert r.status_code == 303

        obj = AttachmentView._get(1)
        assert obj is not None
        assert obj.document is not None
        key = obj.document["key"]

        r = _export(client, "attachment", fmt="csv")
        assert r.status_code == 200
        assert r.headers["content-type"] == "application/zip"

        with zipfile.ZipFile(io.BytesIO(r.content)) as zf:
            names = set(zf.namelist())
            assert "attachment.csv" in names
            asset_path = f"assets/{storage.name}/{key}"
            assert asset_path in names
            assert zf.read(asset_path) == b"report content"

    def test_csv_import_ignores_file_field_column(self, attachment_client):
        """FileField is always excluded from import: a `Document` column in the
        upload is ignored, and the created record's file field stays empty."""
        client, _storage = attachment_client
        upload = _csv_upload([["Name", "Document"], ["Imported", "assets/local/x.txt"]])
        r = client.post(
            "/admin/_api/attachment/import",
            data={"format": "csv"},
            files={"file": upload},
        )
        assert r.status_code == 200
        data = r.json()
        assert data["rows_total"] == 1
        assert data["rows_created"] == 1
        assert not data["has_errors"]

        obj = next(
            (AttachmentView._get(pk) for pk in AttachmentView._all_pks()),
            None,
        )
        assert obj is not None
        assert obj.name == "Imported"
        assert obj.document is None


# ── List page: UI elements ─────────────────────────────────────────────────────


class TestListPageUI:
    def setup_method(self, method):
        ProductView._db.truncate()

    def test_export_button_present_when_can_export(self, client_with_data):
        r = client_with_data.get("/admin/product/list")
        assert r.status_code == 200
        assert "Export" in r.text
        assert 'data-name="export"' in r.text

    def test_import_button_present_when_can_import(self, client_with_data):
        r = client_with_data.get("/admin/product/list")
        assert r.status_code == 200
        assert "Import" in r.text
        assert "modal-import" in r.text

    def test_export_present_no_import_when_disabled(self, export_only_client):
        r = export_only_client.get("/admin/export-only/list")
        assert r.status_code == 200
        assert "Export" in r.text
        assert "modal-import" not in r.text

    def test_no_export_button_when_can_export_false(self, no_export_client):
        r = no_export_client.get("/admin/no-export/list")
        assert r.status_code == 200
        # Neither the export button nor the import modal should appear
        assert 'data-name="export"' not in r.text
        assert "modal-import" not in r.text


# ── Security: import upload-size cap ────────────────────────────────────────────


class TestImportSecurityCaps:
    """Verify that ImportConfig limits are enforced by the import endpoint."""

    def setup_method(self, method):
        ProductView._db.truncate()

    def _make_client(self, **import_kwargs):
        from starlette_admin.importers import ImportConfig

        app = Starlette()
        admin = BaseAdmin(import_config=ImportConfig(**import_kwargs))
        admin.add_view(ProductView(Product))
        admin.mount_to(app)
        return CsrfTestClient(app, raise_server_exceptions=True)

    def test_upload_within_size_limit_is_accepted(self):
        client = self._make_client(max_upload_size=1024 * 1024)
        upload = _csv_upload([["Name", "Price", "Sku"], ["Widget", "9.99", "W-001"]])
        r = client.post(
            "/admin/_api/product/import",
            data={"format": "csv"},
            files={"file": upload},
        )
        assert r.status_code == 200

    def test_upload_exceeding_size_limit_returns_400(self):
        client = self._make_client(max_upload_size=10)  # 10 bytes: tiny cap
        upload = _csv_upload([["Name", "Price", "Sku"], ["Widget", "9.99", "W-001"]])
        r = client.post(
            "/admin/_api/product/import",
            data={"format": "csv"},
            files={"file": upload},
        )
        assert r.status_code == 400
        assert "too large" in r.json().get("error", "").lower()

    @staticmethod
    def _rows(count: int) -> list[list[str]]:
        header = [["Name", "Price", "Sku"]]
        return header + [[f"Item{i}", "1.00", f"S-{i:03d}"] for i in range(count)]

    def test_import_within_row_cap_is_accepted(self):
        client = self._make_client(max_rows=5)
        r = client.post(
            "/admin/_api/product/import",
            data={"format": "csv"},
            files={"file": _csv_upload(self._rows(5))},
        )
        assert r.status_code == 200
        assert r.json()["rows_created"] == 5

    def test_import_exceeding_row_cap_returns_400(self):
        client = self._make_client(max_rows=3)
        r = client.post(
            "/admin/_api/product/import",
            data={"format": "csv"},
            files={"file": _csv_upload(self._rows(4))},
        )
        assert r.status_code == 400
        assert "exceeds the configured maximum" in r.json().get("error", "")
        # The cap fires before any record is created
        assert len(ProductView._db.all()) == 0

    def test_import_row_cap_applies_to_preview(self):
        client = self._make_client(max_rows=3)
        r = client.post(
            "/admin/_api/product/import?preview=1",
            data={"format": "csv"},
            files={"file": _csv_upload(self._rows(4))},
        )
        assert r.status_code == 400

    def test_import_row_cap_disabled_accepts_any_count(self):
        client = self._make_client(max_rows=None)
        r = client.post(
            "/admin/_api/product/import",
            data={"format": "csv"},
            files={"file": _csv_upload(self._rows(50))},
        )
        assert r.status_code == 200
        assert r.json()["rows_created"] == 50

    def test_import_parse_error_with_row_cap_disabled_returns_400(self):
        """With the row cap off, the pre-pass that counts rows (and would
        otherwise surface a parse error) never runs; malformed input is
        instead caught by the generic handler around the main import pass."""
        client = self._make_client(max_rows=None)
        upload = _file_tuple("data.json", b"not valid json", "application/json")
        r = client.post(
            "/admin/_api/product/import",
            data={"format": "json"},
            files={"file": upload},
        )
        assert r.status_code == 400
        assert "error" in r.json()


# ── Security: export row cap ───────────────────────────────────────────────────


class TestExportRowCap:
    """Verify that ExportConfig.max_rows is enforced by the export endpoint."""

    def setup_method(self, method):
        for view in (ProductView, PaginatedProductView):
            view._db.truncate()
            for i in range(5):
                view._db.insert(
                    Product(
                        name=f"Item{i}", price=float(i), sku=f"S-{i:03d}"
                    ).to_tinydb_doc()
                )

    def _make_client(self, max_rows, view_cls=ProductView):
        from starlette_admin.export import ExportConfig

        app = Starlette()
        admin = BaseAdmin(export_config=ExportConfig(max_rows=max_rows))
        admin.add_view(view_cls(Product))
        admin.mount_to(app)
        return CsrfTestClient(app, raise_server_exceptions=True)

    def test_export_within_row_cap_returns_200(self):
        client = self._make_client(max_rows=10)
        r = _export(client, "product", fmt="csv")
        assert r.status_code == 200

    def test_export_exceeding_row_cap_flashes_error(self):
        client = self._make_client(max_rows=3)  # table has 5 rows
        r = _export(client, "product", fmt="csv", follow_redirects=False)
        assert r.status_code == 303

        r = client.get(r.headers["location"], follow_redirects=True)
        assert r.status_code == 200
        assert "exceeds the configured maximum" in r.text

    def test_export_row_cap_disabled_returns_200(self):
        client = self._make_client(max_rows=None)
        r = _export(client, "product", fmt="csv")
        assert r.status_code == 200

    def test_select_all_matching_exceeding_cap_flashes_error(self):
        client = self._make_client(max_rows=3)  # table has 5 rows
        r = _export(
            client,
            "product",
            fmt="csv",
            scope="selected",
            select_all=True,
            follow_redirects=False,
        )
        assert r.status_code == 303

    def test_explicit_selection_exceeding_cap_flashes_error(self):
        client = self._make_client(max_rows=2)
        pks = ProductView._all_pks()  # 5 pks
        r = _export(
            client,
            "product",
            fmt="csv",
            scope="selected",
            pks=pks,
            follow_redirects=False,
        )
        assert r.status_code == 303

    def test_page_scope_within_cap_returns_200(self):
        # Table holds 5 rows, above the cap, but a single page of 2 is fine.
        client = self._make_client(max_rows=3, view_cls=PaginatedProductView)
        r = _export(
            client,
            "paginated-product",
            fmt="csv",
            scope="page",
            list_query="page=1&page_size=2",
        )
        assert r.status_code == 200

    def test_page_scope_partial_last_page_within_cap(self):
        # Page 3 with page_size=2 holds only the fifth row, so the effective
        # export size is 1 even though page_size exceeds the cap of 1.
        client = self._make_client(max_rows=1, view_cls=PaginatedProductView)
        r = _export(
            client,
            "paginated-product",
            fmt="csv",
            scope="page",
            list_query="page=3&page_size=2",
        )
        assert r.status_code == 200

    def test_page_scope_exceeding_cap_flashes_error(self):
        client = self._make_client(max_rows=2, view_cls=PaginatedProductView)
        r = _export(
            client,
            "paginated-product",
            fmt="csv",
            scope="page",
            list_query="page=1&page_size=3",
            follow_redirects=False,
        )
        assert r.status_code == 303

        r = client.get(r.headers["location"], follow_redirects=True)
        assert r.status_code == 200
        assert "exceeds the configured maximum" in r.text

    def test_page_scope_show_all_page_size_cannot_bypass_cap(self):
        # page_size=-1 renders every row, so scope=page must not sidestep
        # the cap when the full table exceeds it.
        client = self._make_client(max_rows=3, view_cls=PaginatedProductView)
        r = _export(
            client,
            "paginated-product",
            fmt="csv",
            scope="page",
            list_query="page=1&page_size=-1",
            follow_redirects=False,
        )
        assert r.status_code == 303

    def test_row_cap_redirect_honors_origin(self):
        client = self._make_client(max_rows=3)  # table has 5 rows
        origin = "/admin/product/list?q=Item&sort=name__asc"
        # "Item" matches all 5 rows, so the cap is still exceeded post-filter.
        r = _export(
            client,
            "product",
            fmt="csv",
            list_query="q=Item&sort=name__asc",
            origin=origin,
            follow_redirects=False,
        )
        assert r.status_code == 303
        assert r.headers["location"] == origin


# ── exclude_fields_from_export / exclude_fields_from_import ───────────────────


class TestExcludeFieldsFromExport:
    def test_excluded_field_absent_from_csv_header(self, exclude_export_client):
        r = _export(exclude_export_client, "exclude-export", fmt="csv")
        assert r.status_code == 200
        reader = csv.reader(io.StringIO(r.text))
        header = next(reader)
        assert "Sku" not in header
        assert "sku" not in header

    def test_non_excluded_fields_present_in_csv_header(self, exclude_export_client):
        r = _export(exclude_export_client, "exclude-export", fmt="csv")
        assert r.status_code == 200
        reader = csv.reader(io.StringIO(r.text))
        header = next(reader)
        assert "Name" in header
        assert "Price" in header

    def test_excluded_field_absent_from_json_export(self, exclude_export_client):
        r = _export(exclude_export_client, "exclude-export", fmt="json")
        assert r.status_code == 200
        rows = r.json()
        assert len(rows) == 1
        keys = set(rows[0].keys())
        assert not any(k.lower() == "sku" for k in keys)

    def test_non_excluded_fields_present_in_json_export(self, exclude_export_client):
        r = _export(exclude_export_client, "exclude-export", fmt="json")
        assert r.status_code == 200
        row = r.json()[0]
        assert row.get("Name") == "Widget" or row.get("name") == "Widget"


class TestExcludeFieldsFromImport:
    def setup_method(self, method):
        ExcludeImportFieldsView._db.truncate()

    def test_excluded_field_is_ignored_during_import(self, exclude_import_client):
        # Upload a CSV that provides price=99.99; it should be ignored and stay 0.0
        upload = _csv_upload([["Name", "Price", "Sku"], ["TestItem", "99.99", "T-001"]])
        r = exclude_import_client.post(
            "/admin/_api/exclude-import/import",
            data={"format": "csv"},
            files={"file": upload},
        )
        assert r.status_code == 200
        data = r.json()
        assert data["rows_created"] == 1

        pk = ExcludeImportFieldsView._all_pks()[0]
        record = ExcludeImportFieldsView._get(pk)
        assert record.price == 0.0  # default: the uploaded value was discarded

    def test_non_excluded_fields_are_still_imported(self, exclude_import_client):
        upload = _csv_upload([["Name", "Price", "Sku"], ["GoodItem", "5.0", "G-001"]])
        r = exclude_import_client.post(
            "/admin/_api/exclude-import/import",
            data={"format": "csv"},
            files={"file": upload},
        )
        assert r.status_code == 200
        pk = ExcludeImportFieldsView._all_pks()[0]
        record = ExcludeImportFieldsView._get(pk)
        assert record.name == "GoodItem"
        assert record.sku == "G-001"
