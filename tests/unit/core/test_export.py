"""Unit tests for starlette_admin.export: base machinery and format exporters."""

from __future__ import annotations

import csv
import io
import json
import re
import zipfile
from dataclasses import dataclass
from typing import Any
from unittest.mock import MagicMock

import openpyxl
import pytest
from starlette.applications import Starlette
from starlette.requests import Request
from starlette.testclient import TestClient
from starlette_admin.exceptions import ExportError
from starlette_admin.export.base import BaseExporter, ExportConfig, ExportContext
from starlette_admin.export.csv import CsvExporter
from starlette_admin.export.helpers import (
    escape_formula,
    is_allowed_url,
    is_file_dict,
    resolve_download_url,
    safe_zip_key,
    safe_zip_name,
)
from starlette_admin.export.json import JsonExporter
from starlette_admin.export.pdf import PdfExporter
from starlette_admin.export.tablib import TablibExporter
from starlette_admin.fields import BaseField, FileField
from starlette_admin.storage.base import BaseStorage

from tests.conftest import make_upload

# ── Helpers ────────────────────────────────────────────────────────────────────


@dataclass
class _Field:
    """Minimal field stub: just name + label."""

    name: str
    label: str | None = None


def _make_file_field(name: str, storage_name: str = "mock") -> FileField:

    class _MockStorage(BaseStorage):
        name = storage_name

        async def save(self, upload, dest):  # type: ignore[override]
            raise NotImplementedError

        async def url(self, request, key, *, signed=False, expires=3600):  # type: ignore[override]
            return f"http://localhost/{key}"

        async def delete(self, key):  # type: ignore[override]
            pass

        async def read(self, key: str) -> bytes:
            return b"file-content-for-" + key.encode()

    storage = _MockStorage()
    field = FileField(name=name)
    field.label = name.replace("_", " ").capitalize()
    field.storage = storage
    return field


def _plain_field(name: str, label: str | None = None) -> BaseField:
    f = BaseField(name=name)
    f.label = label or name.replace("_", " ").capitalize()
    return f


def _make_request() -> Request:
    app = Starlette()
    TestClient(app, raise_server_exceptions=False)
    scope = {
        "type": "http",
        "method": "GET",
        "path": "/",
        "query_string": b"",
        "headers": [(b"host", b"testserver")],
        "server": ("testserver", 80),
        "scheme": "http",
    }
    return Request(scope)


def _make_ctx(
    fields: list[Any],
    rows: list[dict[str, Any]],
    filename: str = "export",
) -> ExportContext:
    return ExportContext(
        fields=fields,
        rows=rows,
        view=MagicMock(),
        request=_make_request(),
        filename=filename,
    )


# ── is_file_dict ─────────────────────────────────────────────────────────────


def test_is_file_dict_true():
    assert is_file_dict({"storage": "local", "key": "a/b.jpg"})


def test_is_file_dict_false_missing_key():
    assert not is_file_dict({"storage": "local"})


def test_is_file_dict_false_not_dict():
    assert not is_file_dict("string")
    assert not is_file_dict(None)
    assert not is_file_dict(42)


# ── _preprocess_rows ──────────────────────────────────────────────────────────


class _ConcreteExporter(BaseExporter):
    content_type = "text/plain"
    extension = "txt"

    async def generate(self, fields, rows):
        return b""


@pytest.fixture()
def exporter() -> _ConcreteExporter:
    return _ConcreteExporter()


def test_preprocess_rows_no_files(exporter):
    fields = [_plain_field("name"), _plain_field("age")]
    rows = [
        {"name": "Alice", "age": 30, "_meta": "ignored"},
        {"name": "Bob", "age": 25},
    ]
    cleaned, file_map, url_file_map = exporter._preprocess_rows(fields, rows)

    assert file_map == {}
    assert url_file_map == {}
    assert cleaned == [{"name": "Alice", "age": 30}, {"name": "Bob", "age": 25}]


def test_preprocess_rows_with_file_field(exporter):
    ff = _make_file_field("avatar", storage_name="mock-store")
    fields = [_plain_field("name"), ff]
    file_value = {
        "storage": "mock-store",
        "key": "avatars/alice.jpg",
        "filename": "alice.jpg",
    }
    rows = [{"name": "Alice", "avatar": file_value}]

    cleaned, file_map, _url_file_map = exporter._preprocess_rows(fields, rows)

    expected_zip_path = "assets/mock-store/avatars/alice.jpg"
    assert cleaned[0]["avatar"] == expected_zip_path
    assert file_map == {expected_zip_path: ("mock-store", "avatars/alice.jpg")}


def test_preprocess_rows_multiple_files(exporter):
    ff = _make_file_field("docs", storage_name="s3")
    ff.multiple = True
    fields = [ff]
    v1 = {"storage": "s3", "key": "docs/a.pdf"}
    v2 = {"storage": "s3", "key": "docs/b.pdf"}
    rows = [{"docs": [v1, v2]}]

    cleaned, file_map, _url_file_map = exporter._preprocess_rows(fields, rows)

    val = cleaned[0]["docs"]
    assert "assets/s3/docs/a.pdf" in val
    assert "assets/s3/docs/b.pdf" in val
    assert len(file_map) == 2


def test_preprocess_rows_none_file_value(exporter):
    ff = _make_file_field("avatar")
    fields = [ff]
    rows = [{"avatar": None}]
    cleaned, file_map, _url_file_map = exporter._preprocess_rows(fields, rows)
    assert cleaned[0]["avatar"] == ""
    assert file_map == {}


# ── _build_zip ────────────────────────────────────────────────────────────────


def test_build_zip_contains_format_and_assets(exporter):
    asset_files = {"assets/local/img.png": b"\x89PNG..."}
    zip_bytes = exporter._build_zip(b"data,col\n", "export.csv", asset_files)

    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        names = set(zf.namelist())
        assert "export.csv" in names
        assert "assets/local/img.png" in names
        assert zf.read("export.csv") == b"data,col\n"
        assert zf.read("assets/local/img.png") == b"\x89PNG..."


# ── build_response: no-file path ──────────────────────────────────────────────


@pytest.mark.asyncio
async def test_build_response_no_files_returns_direct():
    exporter = CsvExporter()
    fields = [_plain_field("name"), _plain_field("age")]
    ctx = _make_ctx(fields, [{"name": "Alice", "age": 30}])

    resp = await exporter.build_response(ctx)

    assert resp.media_type == "text/csv"
    assert "export.csv" in resp.headers["content-disposition"]
    body = b"".join(resp.body_iterator) if hasattr(resp, "body_iterator") else resp.body
    text = body.decode()
    assert "Name" in text
    assert "Alice" in text


# ── build_response: file-field ZIP path ───────────────────────────────────────


@pytest.mark.asyncio
async def test_build_response_with_file_field_returns_zip(storage):
    info = await storage.save(
        make_upload(b"file content", "alice.jpg"), "photos/alice.jpg"
    )
    ff = FileField(name="photo", storage=storage)
    fields = [_plain_field("name"), ff]
    file_value = {
        "storage": storage.name,
        "key": info.key,
        "filename": "alice.jpg",
    }
    ctx = _make_ctx(fields, [{"name": "Alice", "photo": file_value}])

    exporter = CsvExporter()
    resp = await exporter.build_response(ctx)

    assert resp.media_type == "application/zip"
    assert "export.zip" in resp.headers["content-disposition"]
    body = b"".join(resp.body_iterator) if hasattr(resp, "body_iterator") else resp.body

    with zipfile.ZipFile(io.BytesIO(body)) as zf:
        names = set(zf.namelist())
        assert "export.csv" in names
        asset_path = f"assets/{storage.name}/{info.key}"
        assert asset_path in names
        assert zf.read(asset_path) == b"file content"


# ── CsvExporter ───────────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_csv_generate_headers_and_rows():
    exporter = CsvExporter()
    fields = [_plain_field("name", "Full Name"), _plain_field("age", "Age")]
    rows = [{"name": "Alice", "age": 30}, {"name": "Bob", "age": ""}]

    result = await exporter.generate(fields, rows)
    text = result.decode("utf-8")
    reader = csv.reader(io.StringIO(text))
    lines = list(reader)

    assert lines[0] == ["Full Name", "Age"]
    assert lines[1] == ["Alice", "30"]
    assert lines[2] == ["Bob", ""]


@pytest.mark.asyncio
async def test_csv_generate_uses_field_name_when_no_label():
    exporter = CsvExporter()
    f = BaseField(name="my_field")
    f.label = None
    result = await exporter.generate([f], [{"my_field": "val"}])
    assert b"my_field" in result


@pytest.mark.asyncio
async def test_csv_generate_empty_rows():
    exporter = CsvExporter()
    fields = [_plain_field("x")]
    result = await exporter.generate(fields, [])
    lines = result.decode().strip().splitlines()
    assert len(lines) == 1  # header only


# ── escape_formula ───────────────────────────────────────────────────────────


@pytest.mark.parametrize("trigger", ["=", "+", "-", "@"])
def test_escape_formula_prefixes_trigger_chars(trigger):
    assert escape_formula(f"{trigger}cmd|'/c calc'!A1") == f"'{trigger}cmd|'/c calc'!A1"


def test_escape_formula_leaves_safe_value_untouched():
    assert escape_formula("Alice") == "Alice"


def test_escape_formula_empty_string():
    assert escape_formula("") == ""


@pytest.mark.parametrize("prefix", [" ", "  ", "\n", "\n\n", "\x0b", "\x0c"])
@pytest.mark.parametrize("trigger", ["=", "+", "-", "@"])
def test_escape_formula_prefixes_after_leading_whitespace(prefix, trigger):
    value = f"{prefix}{trigger}cmd|'/c calc'!A1"
    assert escape_formula(value) == f"'{value}"


@pytest.mark.parametrize("trigger", ["\t", "\r"])
def test_escape_formula_prefixes_tab_and_cr(trigger):
    value = f"{trigger}not even a trigger char follows"
    assert escape_formula(value) == f"'{value}"


def test_escape_formula_leaves_whitespace_only_value_untouched():
    assert escape_formula("   ") == "   "


# ── TablibExporter (xlsx) ────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_excel_generate_headers_and_rows():

    exporter = TablibExporter("xlsx")
    fields = [_plain_field("name", "Full Name"), _plain_field("score", "Score")]
    rows = [{"name": "Alice", "score": 95}, {"name": "Bob", "score": 82}]

    result = await exporter.generate(fields, rows)
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb.active
    data = list(ws.iter_rows(values_only=True))

    assert data[0] == ("Full Name", "Score")
    assert data[1] == ("Alice", 95)
    assert data[2] == ("Bob", 82)


@pytest.mark.asyncio
async def test_excel_generate_empty_rows():

    exporter = TablibExporter("xlsx")
    fields = [_plain_field("x", "X")]
    result = await exporter.generate(fields, [])
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb.active
    data = list(ws.iter_rows(values_only=True))
    assert data == [("X",)]


@pytest.mark.asyncio
async def test_excel_content_type():
    exporter = TablibExporter("xlsx")
    assert "spreadsheetml" in exporter.content_type


# ── JsonExporter ──────────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_json_generate_basic():
    exporter = JsonExporter()
    fields = [_plain_field("name", "Name"), _plain_field("city", "City")]
    rows = [{"name": "Alice", "city": "Paris"}, {"name": "Bob", "city": "Lyon"}]

    result = await exporter.generate(fields, rows)
    data = json.loads(result.decode("utf-8"))

    assert data == [
        {"Name": "Alice", "City": "Paris"},
        {"Name": "Bob", "City": "Lyon"},
    ]


@pytest.mark.asyncio
async def test_json_generate_non_serializable_falls_back_to_str():
    from datetime import date

    exporter = JsonExporter()
    fields = [_plain_field("dt", "Date")]
    rows = [{"dt": date(2024, 1, 1)}]

    result = await exporter.generate(fields, rows)
    data = json.loads(result)
    assert data[0]["Date"] == "2024-01-01"


@pytest.mark.asyncio
async def test_json_generate_empty():
    exporter = JsonExporter()
    result = await exporter.generate([_plain_field("x")], [])
    assert json.loads(result) == []


@pytest.mark.asyncio
async def test_json_generate_missing_value_defaults_empty_string():
    exporter = JsonExporter()
    fields = [_plain_field("a", "A"), _plain_field("b", "B")]
    rows = [{"a": "only-a"}]  # b is missing

    result = await exporter.generate(fields, rows)
    data = json.loads(result)
    assert data[0]["B"] == ""


# ── PdfExporter ───────────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_pdf_generate_returns_pdf_bytes():
    exporter = PdfExporter()
    fields = [_plain_field("name", "Name"), _plain_field("age", "Age")]
    rows = [{"name": "Alice", "age": 30}]

    result = await exporter.generate(fields, rows)
    assert result[:4] == b"%PDF"


@pytest.mark.asyncio
async def test_pdf_generate_empty_rows():
    exporter = PdfExporter()
    fields = [_plain_field("x", "X")]
    result = await exporter.generate(fields, [])
    assert result[:4] == b"%PDF"


@pytest.mark.asyncio
async def test_pdf_generate_no_fields():
    exporter = PdfExporter()
    # Edge case: zero fields should not crash
    result = await exporter.generate([], [])
    assert result[:4] == b"%PDF"


# ── _zip_path_for_url_file ────────────────────────────────────────────────────


def test_zip_path_for_url_file_without_filename():
    exporter = _ConcreteExporter()
    url_file_map: dict[str, str] = {}
    zip_path = exporter._zip_path_for_url_file(
        {"url": "https://example.com/path/photo.png"}, url_file_map
    )
    assert zip_path.startswith("assets/url/")
    assert zip_path.endswith("_photo.png")
    assert url_file_map[zip_path] == "https://example.com/path/photo.png"


# ── _export_file_value fallback ───────────────────────────────────────────────


def test_export_file_value_dict_without_storage_key_or_url():
    exporter = _ConcreteExporter()
    assert exporter._export_file_value({"filename": "x.txt"}, {}, {}) == "x.txt"


def test_export_file_value_non_dict_value():
    exporter = _ConcreteExporter()
    assert exporter._export_file_value("raw", {}, {}) == "raw"
    assert exporter._export_file_value(None, {}, {}) == ""


# ── _preprocess_rows: file fields without storage ─────────────────────────────


def test_preprocess_rows_single_file_without_storage_passthrough(exporter):
    ff = FileField(name="legacy")
    fields = [ff]
    rows = [{"legacy": {"url": "https://example.com/f.txt"}}]
    cleaned, file_map, _url_file_map = exporter._preprocess_rows(fields, rows)
    assert re.compile(r"^assets/url/[^/]*f\.txt$").match(cleaned[0]["legacy"])
    assert file_map == {}


def test_preprocess_rows_multiple_files_without_storage_passthrough(exporter):
    ff = FileField(name="legacy", multiple=True)
    fields = [ff]
    rows = [
        {
            "legacy": [
                {"url": "https://example.com/a.txt"},
                {"url": "https://example.com/b.txt"},
            ]
        },
    ]
    cleaned, file_map, _url_file_map = exporter._preprocess_rows(fields, rows)

    assert re.compile(r"^assets/url/[^/]*a\.txt\nassets/url/[^/]*b\.txt$").match(
        cleaned[0]["legacy"]
    )
    assert file_map == {}


# ── _fetch_files: skip missing / unreadable storage ───────────────────────────


@pytest.mark.asyncio
async def test_fetch_files_skips_unreadable_storage(exporter):
    from starlette_admin.storage.base import BaseStorage, register_storage

    class _FailingStorage(BaseStorage):
        name = "failing-store"

        async def save(self, upload, dest):  # type: ignore[override]
            raise NotImplementedError

        async def url(self, request, key, *, signed=False, expires=3600):  # type: ignore[override]
            return ""

        async def delete(self, key):  # type: ignore[override]
            pass

        async def read(self, key: str) -> bytes:
            raise FileNotFoundError(key)

    register_storage(_FailingStorage())
    fetched = await exporter._fetch_files(
        {"assets/failing-store/x.txt": ("failing-store", "x.txt")}
    )
    assert fetched == {}


# ── _fetch_url_files ──────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_fetch_url_files_downloads_successfully(exporter):
    from unittest.mock import patch

    class _FakeResponse:
        def read(self, n=None) -> bytes:
            return b"downloaded"

        def __enter__(self):
            return self

        def __exit__(self, *args):
            pass

    # restrict_url_download=False so the allow-list is bypassed; we're testing
    # the download mechanics, not the URL restriction logic (covered separately).
    config = ExportConfig(restrict_url_download=False)
    with patch("urllib.request.urlopen", return_value=_FakeResponse()):
        fetched = await exporter._fetch_url_files(
            {"assets/url/abc_file.txt": "https://example.com/file.txt"},
            config,
            _make_request(),
        )
    assert fetched == {"assets/url/abc_file.txt": b"downloaded"}


# ── is_allowed_url ───────────────────────────────────────────────────────────


def test_is_allowed_url_same_origin():
    req = _make_request()
    assert is_allowed_url("http://testserver/uploads/photo.jpg", req)


def test_is_allowed_url_foreign_origin_blocked():
    req = _make_request()
    assert not is_allowed_url("https://evil.com/photo.jpg", req)


def test_is_allowed_url_amazonaws_blocked():
    req = _make_request()
    assert not is_allowed_url("https://s3.amazonaws.com/my-bucket/file.jpg", req)


# ── resolve_download_url ─────────────────────────────────────────────────────


def test_resolve_download_url_restrict_blocks_foreign():
    config = ExportConfig(restrict_url_download=True)
    url = "https://cdn.example.com/photo.jpg"
    assert resolve_download_url(url, config, _make_request()) is None


def test_resolve_download_url_unrestricted_allows_any():
    config = ExportConfig(restrict_url_download=False)
    url = "https://cdn.example.com/photo.jpg"
    assert resolve_download_url(url, config, _make_request()) == url


def test_resolve_download_url_custom_callback_overrides():
    signed = "https://cdn.example.com/bucket/file.jpg?signed=1"
    config = ExportConfig(
        restrict_url_download=True,
        safe_download_url=lambda url, req: signed,
    )
    url = "https://cdn.example.com/photo.jpg"
    assert resolve_download_url(url, config, _make_request()) == signed


def test_resolve_download_url_custom_callback_can_block():
    config = ExportConfig(
        restrict_url_download=False,
        safe_download_url=lambda url, req: None,
    )
    assert (
        resolve_download_url("https://cdn.example.com/b/f", config, _make_request())
        is None
    )


# ── _fetch_url_files: URL blocking ────────────────────────────────────────────


@pytest.mark.asyncio
async def test_fetch_url_files_blocks_foreign_url_by_default(exporter):
    config = ExportConfig(restrict_url_download=True)
    fetched = await exporter._fetch_url_files(
        {"assets/url/abc_photo.jpg": "https://cdn.example.com/photo.jpg"},
        config,
        _make_request(),
    )
    assert fetched == {}


@pytest.mark.asyncio
async def test_fetch_url_files_allows_same_origin(exporter):
    from unittest.mock import patch

    class _FakeResponse:
        def read(self, n=None) -> bytes:
            return b"img"

        def __enter__(self):
            return self

        def __exit__(self, *args):
            pass

    config = ExportConfig(restrict_url_download=True)
    req = _make_request()
    base = str(req.base_url).rstrip("/")
    with patch("urllib.request.urlopen", return_value=_FakeResponse()):
        fetched = await exporter._fetch_url_files(
            {"assets/url/abc_img.png": f"{base}/uploads/img.png"},
            config,
            req,
        )
    assert fetched == {"assets/url/abc_img.png": b"img"}


# ── max_download_size ─────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_fetch_url_files_skips_oversized_download(exporter):
    from unittest.mock import patch

    class _BigResponse:
        def read(self, n=None):
            # Always return more than the limit
            return b"x" * (n or 100)

        def __enter__(self):
            return self

        def __exit__(self, *args):
            pass

    config = ExportConfig(restrict_url_download=False, max_download_size=5)
    with patch("urllib.request.urlopen", return_value=_BigResponse()):
        fetched = await exporter._fetch_url_files(
            {"assets/url/abc_big.bin": "https://example.com/big.bin"},
            config,
            _make_request(),
        )
    assert fetched == {}


@pytest.mark.asyncio
async def test_fetch_url_files_accepts_within_size_limit(exporter):
    from unittest.mock import patch

    class _SmallResponse:
        def read(self, n=None):
            return b"hi"

        def __enter__(self):
            return self

        def __exit__(self, *args):
            pass

    config = ExportConfig(restrict_url_download=False, max_download_size=10)
    with patch("urllib.request.urlopen", return_value=_SmallResponse()):
        fetched = await exporter._fetch_url_files(
            {"assets/url/abc_small.txt": "https://example.com/small.txt"},
            config,
            _make_request(),
        )
    assert fetched == {"assets/url/abc_small.txt": b"hi"}


@pytest.mark.asyncio
async def test_fetch_url_files_no_size_limit(exporter):
    from unittest.mock import patch

    class _UnboundedResponse:
        def read(self, n=None) -> bytes:
            return b"unlimited content"

        def __enter__(self):
            return self

        def __exit__(self, *args):
            pass

    config = ExportConfig(restrict_url_download=False, max_download_size=None)
    with patch("urllib.request.urlopen", return_value=_UnboundedResponse()):
        fetched = await exporter._fetch_url_files(
            {"assets/url/abc_unlimited.bin": "https://example.com/unlimited.bin"},
            config,
            _make_request(),
        )
    assert fetched == {"assets/url/abc_unlimited.bin": b"unlimited content"}


# ── TablibExporter non-native cell values (tablib stringifies via str()) ───────


@pytest.mark.asyncio
async def test_excel_generate_list_value_becomes_str():
    exporter = TablibExporter("xlsx")
    fields = [_plain_field("tags", "Tags")]
    rows = [{"tags": ["a", "b", "c"]}]
    result = await exporter.generate(fields, rows)
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb.active
    data = list(ws.iter_rows(values_only=True))
    assert data[1][0] == str(["a", "b", "c"])


@pytest.mark.asyncio
async def test_excel_generate_tuple_value_becomes_str():
    exporter = TablibExporter("xlsx")
    fields = [_plain_field("coords", "Coords")]
    rows = [{"coords": (1, 2, 3)}]
    result = await exporter.generate(fields, rows)
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb.active
    data = list(ws.iter_rows(values_only=True))
    assert data[1][0] == str((1, 2, 3))


# ── ZIP path sanitization ─────────────────────────────────────────────────────


def test_safe_zip_name_strips_directory_prefix():
    assert safe_zip_name("../evil", "storage_name") == "evil"


def test_safe_zip_name_rejects_dotdot_literal():
    with pytest.raises(ExportError, match="Unsafe ZIP storage_name"):
        safe_zip_name("..", "storage_name")


def test_safe_zip_name_rejects_backslash_traversal():
    # backslash normalized to slash before basename
    assert safe_zip_name("..\\evil", "filename") == "evil"


def test_safe_zip_name_empty_becomes_underscore():
    assert safe_zip_name("", "filename") == "_"


def test_safe_zip_key_accepts_normal_path():
    assert safe_zip_key("covers/photo.jpg") == "covers/photo.jpg"


def test_safe_zip_key_rejects_dotdot_segment():
    with pytest.raises(ExportError, match="Unsafe ZIP key contains"):
        safe_zip_key("covers/../../../etc/passwd")


def test_safe_zip_key_rejects_leading_dotdot():
    with pytest.raises(ExportError, match="Unsafe ZIP key contains"):
        safe_zip_key("../secret.txt")


def test_zip_path_for_file_rejects_dotdot_in_storage_name(exporter):
    with pytest.raises(ExportError):
        exporter._zip_path_for_file({"storage": "..", "key": "file.txt"}, {})


def test_zip_path_for_file_rejects_dotdot_in_key(exporter):
    with pytest.raises(ExportError):
        exporter._zip_path_for_file(
            {"storage": "local", "key": "uploads/../../../etc/passwd"}, {}
        )


def test_zip_path_for_url_file_rejects_dotdot_filename(exporter):
    with pytest.raises(ExportError):
        exporter._zip_path_for_url_file({"url": "http://x.com/f", "filename": ".."}, {})


@pytest.mark.asyncio
async def test_excel_generate_custom_object_becomes_str():
    exporter = TablibExporter("xlsx")
    fields = [_plain_field("meta", "Meta")]
    rows = [{"meta": {"key": "value"}}]  # dict is not an Excel-native type
    result = await exporter.generate(fields, rows)
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb.active
    data = list(ws.iter_rows(values_only=True))
    assert data[1][0] == str({"key": "value"})


# ── Scheme restriction (SSRF guard) ──────────────────────────────────────────


@pytest.mark.asyncio
@pytest.mark.parametrize(
    "bad_url",
    [
        "file:///etc/passwd",
        "file://localhost/etc/shadow",
        "ftp://internal.host/secret.txt",
        "gopher://evil.com/",
        "dict://localhost:11211/stats",
    ],
)
async def test_fetch_url_files_blocks_non_http_schemes(exporter, bad_url):
    """Non-http(s) URLs must be blocked even when restrict_url_download=False.

    urllib.request.urlopen supports file://, ftp://, etc., and allowing them
    would be an SSRF vector (local file read, internal service probe).
    """
    config = ExportConfig(restrict_url_download=False)
    fetched = await exporter._fetch_url_files(
        {"assets/url/test_file": bad_url},
        config,
        _make_request(),
    )
    assert fetched == {}, f"Expected {bad_url!r} to be blocked but it was fetched"


# ── Content-Disposition filename sanitization ─────────────────────────────────


@pytest.mark.asyncio
async def test_build_response_sanitizes_filename_in_content_disposition():
    """Filenames with injected quote/newline chars must not leak into the header."""
    exporter = CsvExporter()
    ctx = _make_ctx(
        [_plain_field("x")],
        [{"x": "val"}],
        filename='evil"name\r\nX-Injected: header',
    )
    resp = await exporter.build_response(ctx)
    cd = resp.headers.get("content-disposition", "")
    assert '"' not in cd.split("filename=", 1)[-1] or cd.count('"') == 2
    assert "\r" not in cd
    assert "\n" not in cd
